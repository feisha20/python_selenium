#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import os
import configparser
import selenium.webdriver
import selenium.webdriver.common.action_chains
import selenium.webdriver.common.keys
import time
import find
import openpyxl

# 项目的基本信息
projectName = "pytest_selenium"
configFileName = "QA.conf"
file_path = os.path.abspath(os.getcwd())
index = int(file_path.index(projectName))
project_path = file_path[0:index] + projectName
config_path = project_path + "\\conf\\"
config_file_path = config_path + configFileName

# 基本配置
load_config = False
cp = configparser.ConfigParser()


# 加载配置
def is_load():
    return load_config


# 获取配置文件
def get_config(f=config_file_path):
    global load_config
    if not f == config_file_path:
        f = config_path + f
        cp.read(f, encoding="utf-8-sig")
    else:
        if not is_load():
            cp.read(f, encoding="utf-8-sig")
            load_config = True
    return cp


# 获取配置文件某项的信息
def get_config_item(section, option, f=None):
    global load_config
    if not f:
        f = file_path
        if not is_load():
            cp.read(f, encoding="utf-8-sig")
            load_config = True
    else:
        f = config_path + f
        cp.read(f, encoding="utf-8-sig")
    value = str(cp.get(section, option))
    return value


# 写入接收邮件的配置信息：
def write_mail_conf(mail):
    conf_file = config_path + "MAIL.conf"
    with open(conf_file, 'w') as f:
        f.write('[mail_report]\nreceive_mail=' + mail)


selenium_browser = None


# 启动浏览器
def start_browser():
    browser = get_config_item("browser", "browser", "BR.conf")
    global selenium_browser
    if browser == "Chrome":
        kill_chrome_process()
        chrome_driver = get_chromedriver_path()
        print(chrome_driver)
        os.environ["webdriver.chrome.driver"] = chrome_driver
        option = selenium.webdriver.ChromeOptions()
        option.add_argument("--start-maximized")
        option.add_argument("--test-type")
        option.add_argument("--disable-web-security")
        selenium_browser = selenium.webdriver.Chrome(chrome_options=option, executable_path=chrome_driver)
    elif browser == "Firefox":
        kill_firefox_process()
        firefox_driver = get_firefoxdriver_path()
        os.environ["webdriver.firefox.driver"] = firefox_driver
        selenium_browser = selenium.webdriver.Firefox(executable_path=firefox_driver)
    elif browser == "IE":
        kill_ie_process()
        ie_driver = get_iedriver_path()
        os.environ["webdriver.Ie.driver"] = ie_driver
        selenium_browser = selenium.webdriver.Ie(executable_path=ie_driver)
    return selenium_browser


# 打开url
def open_url(url=None):
    start_browser().get(url)


# 获取当前页的url
def get_url():
    url = selenium_browser.current_url
    return url


# 获取当前页的title
def get_title():
    title = selenium_browser.title
    return title


# 释放selenium对象，关闭driver和chrome进程
def stop_chrome():
    selenium_browser.close()
    kill_chrome_process()


def kill_chrome_process():
    """杀掉谷歌进程"""
    killed = False
    # noinspection PyBroadException
    try:
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq chrome.exe"')
            text = str(o.read())
            if "chrome.exe" in text:
                os.popen("taskkill /im chrome.exe /F")
                time.sleep(1)
            else:
                killed = True
        killed = False
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq chromedriver.exe"')
            text = str(o.read())
            if "chromedriver.exe" in text:
                os.popen("taskkill /im chromedriver.exe /F")
                time.sleep(1)
            else:
                killed = True
    except:
        print("进程chrome.exe 或者 chromedriver.exe不存在")


def kill_firefox_process():
    """杀掉Firefox进程"""
    killed = False
    # noinspection PyBroadException
    try:
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq firefox.exe"')
            text = str(o.read())
            if "firefox.exe" in text:
                os.popen("taskkill /im firefox.exe /F")
                time.sleep(1)
            else:
                killed = True
        killed = False
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq geckodriver.exe"')
            text = str(o.read())
            if "geckodriver.exe" in text:
                os.popen("taskkill /im geckodriver.exe /F")
                time.sleep(1)
            else:
                killed = True
    except:
        print("进程firefox.exe 或者 geckodriver.exe不存在")


def kill_ie_process():
    """杀掉IE进程"""
    killed = False
    # noinspection PyBroadException
    try:
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq iexplore.exe"')
            text = str(o.read())
            if "iexplore.exe" in text:
                os.popen("taskkill /im iexplore.exe /F")
                time.sleep(1)
            else:
                killed = True
        killed = False
        while not killed:
            o = os.popen('tasklist /FI "IMAGENAME eq IEDriverServer.exe"')
            text = str(o.read())
            if "IEDriverServer.exe" in text:
                os.popen("taskkill /im IEDriverServer.exe /F")
                time.sleep(1)
            else:
                killed = True
    except:
        print("进程iexplore.exe 或者 IEDriverServer.exe不存在")


# 获取selenium对象
def get_selenium():
    return selenium_browser


# 保持浏览器页面页签始终只有一个
def keep_window():
    get_selenium().close()
    switch_window(0)


# 切换浏览器页签
def switch_window(window_index=0):
    handles = get_selenium().window_handles
    get_selenium().switch_to.window(handles[window_index])


# 获取chromedriver绝对路径
def get_chromedriver_path():
    return str(project_path + "\\driver\\chromedriver.exe")


# 获取firefoxdriver绝对路径
def get_firefoxdriver_path():
    return str(project_path + "\\driver\\geckodriver.exe")


# 获取iedriver绝对路径
def get_iedriver_path():
    return str(project_path + "\\driver\\IEDriverServer.exe")


# 切入iframe，解决获取不到元素的问题
def switch_iframe(iframe=None):
    if not iframe:
        return get_selenium().switch_to.frame(find.element("t,iframe"))
    elif str(iframe).lower() == "default":
        return get_selenium().switch_to.default_content()
    else:
        return get_selenium().switch_to.frame(iframe)


# 切入iframe后，记得返回主文档才能操作主文档的内容
def switch_default():
    get_selenium().switch_to.default_content()


# 返回上一级iframe
def switch_parent_iframe():
    get_selenium().switch_to.parent_frame()


# 定义pytest调试方法参数
def get_pytest_param(file_name, option=None):
    html_file = project_path + "\\report\\report.html"
    html_file = html_file.replace("\\report ", "")
    file_name = file_name + " --html=" + html_file
    if option:
        file_name = file_name + " " + option
    return file_name


def data_file_path(data_file_name):
    """data数据文件的路径"""
    data_file = project_path + "\\data\\" + data_file_name
    return data_file


def get_excelData_by_key(key, data_file_name, sheetname):

    """根据excel第一列的key值查询对应的value值（对应的第二列的值）"""
    data_file = data_file_path(data_file_name)
    wb = openpyxl.load_workbook(data_file)
    ws = wb[sheetname]
    row = 1
    col = 1
    while ws.cell(row=row, column=col).value != key:
        row += 1
    else:
        cc = ws.cell(row=row, column=2).value
    wb.close()
    return cc


def get_excelData_by_cell(row, col, data_file_name, sheetname):

    """根据excel的坐标值取cell data值"""
    data_file = data_file_path(data_file_name)
    wb = openpyxl.load_workbook(data_file)
    ws = wb[sheetname]
    cc = ws.cell(row=row, column=col).value
    if cc is not None:
        return cc
    else:
        print("没有找到数据！")
    wb.close()


def set_run_browser(browser="Chrome", file_name="BR.conf"):
    """设置启动浏览器配置文件"""
    config_file = config_path + file_name
    with open(config_file, 'w') as f:
        f.write('[browser]\nbrowser = ' + browser)
